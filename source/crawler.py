# -*- coding:utf-8 -*-

import getpass
import openpyxl
import os
import random
import re
import time
from bs4 import BeautifulSoup
from datetime import datetime
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException
from urllib.request import urlopen
from urllib.error import URLError, HTTPError


class Crawler:

    def __init__(self):
        self.TIMEOUT_MIN = 0
        self.TIMEOUT_MAX = 1
        self.TIMEOUT_EXPLICIT = 10  # 명시적 타임아웃
        
        self.COLUMN_URL = 'G'  # 게시물 주소
        self.COLUMN_REPLY_COUNT = 'I'  # 댓글수
        self.COLUMN_VIEW_COUNT = 'H'  # 조회수
        
        self.inputFile = None  # 입력 파일
        self.outputFile = None  # 출력 파일
        self.userAccount = {'id': '', 'pw': ''}  # 사용자 계정
        
        self.driver = None
        self.book = None
        self.sheet = None

    # 법적 책임 동의
    def confirmAgreement(self): 
        print("이 프로그램 사용으로 인한 모든 법적 책임은 전적으로 사용자에게 있습니다.\n동의하십니까? (Y/N)")
        confirm = input().strip().upper()
        if not confirm == 'Y':
            return False
        else :
            return True

    # 입력 파일명
    def getInputFileName(self):
        print("\n입력 파일명을 입력해주세요.")
        self.inputFile = input("input file name: ").strip()
    
    # 출력 파일명
    def getOutputFileName(self):
        print("\n출력 파일명을 입력해주세요.")
        self.outputFile = input("output file name: ").strip()
    
    # 네이버 계정
    def getUserAccount(self):
        print("\n계정 정보를 입력해주세요.")
        self.userAccount['id'] = input("naver id: ").strip().lower()
        self.userAccount['pw'] = getpass.getpass(prompt="naver pw: ").strip()
    
    # 파일 존재 유무 확인
    def isExistFile(self, fileName):
        print('\n파일 존재 유무 확인')
        
        fileUrl = os.getcwd() + "/" + self.inputFile + ".xlsx"
        
        if os.path.isfile(fileUrl) and os.path.exists(fileUrl):
            print('파일 확인 완료')
            return True
        else:
            print('파일이 존재하지 않습니다.')
            return False
    
    # 파일 열기
    def openInputFile(self):
        print('\n파일 열기')
        try:
            # 엑셀 파일 읽기
            fileUrl = os.getcwd() + "/" + self.inputFile + ".xlsx"
            
            # data_only=True로 해줘야 수식이 아닌 값으로 받아온다.
            self.book = openpyxl.load_workbook(fileUrl, data_only=True)
            
            # 활성화된 시트 추출하기
            self.sheet = self.book.active
            
            print('파일 열기 완료')
            return True
        except FileNotFoundError as e:
            print("파일이 존재하지 않습니다.\n", e)
            return False
        
    # 크롬 드라이버 연결
    def connectWebDriver(self):
        if not self.driver:
            print('\n크롬 드라이버 연결')
            options = webdriver.ChromeOptions()
            options.add_argument('headless')
            options.add_experimental_option('excludeSwitches', ['enable-logging'])
            self.driver = webdriver.Chrome(os.getcwd() + '/driver/chromedriver.exe', options=options)
        
    # 크롬 드라이버 연결 해제
    def closeWebDriver(self):
        if self.driver:
            print('\n크롬 드라이버 연결 해제')
            self.driver.close()
        
    # 로그인하기
    def login(self):
        print('\n네이버 로그인')
        
        # 네이버 로그인 페이지 접속
        time.sleep(random.randrange(self.TIMEOUT_MIN, self.TIMEOUT_MAX))
        self.driver.get('https://nid.naver.com/nidlogin.login')
        
        try:
            WebDriverWait(self.driver, self.TIMEOUT_EXPLICIT).until(EC.presence_of_element_located((By.ID, 'log.login')))
                    
        finally:
            # driver.maximize_window() # 브라우저 화면 최대화
             
            # 아이디 입력
            time.sleep(random.randrange(self.TIMEOUT_MIN, self.TIMEOUT_MAX))
            self.driver.execute_script("document.getElementsByName('id')[0].value=\'" + self.userAccount['id'] + "\'")
            
            # 비밀번호 입력
            time.sleep(random.randrange(self.TIMEOUT_MIN, self.TIMEOUT_MAX))
            self.driver.execute_script("document.getElementsByName('pw')[0].value=\'" + self.userAccount['pw'] + "\'")
            
            # 로그인 버튼 클릭
            time.sleep(random.randrange(self.TIMEOUT_MIN, self.TIMEOUT_MAX))
            self.driver.find_element_by_class_name('btn_global').click()
                
    # 로그인 확인
    def isLogin(self):
        print('\n네이버 로그인 확인중')
        
        try:
            WebDriverWait(self.driver, self.TIMEOUT_EXPLICIT).until(EC.presence_of_element_located((By.ID, 'content')))
        
        finally:
            time.sleep(self.TIMEOUT_MAX)
            html = self.driver.page_source  # 현재 페이지의 주소를 반환 
            soup = BeautifulSoup(html, 'html.parser')
            
            loginError = soup.find_all(class_='error', attrs={"aria-live": "assertive"})
            if loginError:
                print("가입하지 않은 아이디이거나, 잘못된 비밀번호입니다.")
                return False
            else:
                print('네이버 로그인 완료')
                return True
    
    # 데이터 크롤링
    def getData(self):
        print('\n데이터 크롤링')
        print('입력 파일의 행 개수: ', self.sheet.max_row)
        
        for i in range(1, self.sheet.max_row):
            # 읽기
            print()
            print(i, "번째 행")
            postUrl = self.sheet[self.COLUMN_URL + str(i)].value  # 게시물 주소
            
            # 빈 값인지 확인
            if not postUrl:
                continue
            
            # url 형식에 맞는지 확인
            p = re.compile('(http(s)?:\/\/|www.)([a-z0-9\w]+\.*)+[a-z0-9]{2,4}([\/a-z0-9-%#?&=\w])+(\.[a-z0-9]{2,4}(\?[\/a-z0-9-%#?&=\w]+)*)*')
            match = p.match(postUrl)
            if not match:
                continue
            
            # 존재하는 페이지인지 확인
            try:
                print("게시물 주소:", postUrl)
                res = urlopen(postUrl)
                # print(res.status)
            except ValueError as e:
                print("올바르지 않은 주소 형식입니다.\n", e)
                continue
            except HTTPError as e:
                err = e.read()
                code = e.getcode()
                print(code)  # 404
             
            # 네이버 카페 페이지 들어가기
            try:
                self.driver.get(postUrl)
                replyCount = '삭제된 게시물'  # 댓글수
                viewCount = '삭제된 게시물'  # 조회수
                WebDriverWait(self.driver, self.TIMEOUT_EXPLICIT).until(EC.presence_of_element_located((By.ID, 'cafe_check')))
                 
                try:  # 경고창이 있을 경우
                    WebDriverWait(self.driver, self.TIMEOUT_MAX).until(EC.alert_is_present())
                    alert = self.driver.switch_to.alert
                    print(alert.text)  # 삭제되었거나 없는 게시글입니다.
                    alert.accept()
                     
                except:  # 경고창이 없을 경우
                    WebDriverWait(self.driver, self.TIMEOUT_EXPLICIT).until(EC.presence_of_element_located((By.ID, 'cafe_main')))
                    html = self.driver.page_source  # 현재 페이지의 주소를 반환 
                    soup = BeautifulSoup(html, 'html.parser')
                           
                    # 카페 포스트 본문을 보여주는 iframe 주소를 찾는다
                    iframes = soup.find_all('iframe', id="cafe_main")
                           
                    # for iframe in iframes:
                    #    print(iframe.get('name'))
                           
                    self.driver.switch_to_default_content  # 상위 프레임으로 전환
                    self.driver.switch_to.frame('cafe_main')  # cafe_main 프레임으로 전환
                           
                    html = self.driver.page_source  # 현재 페이지의 주소를 반환 
                    soup = BeautifulSoup(html, 'html.parser')
                     
                    # 댓글수
                    replyEl = soup.find_all(class_='ArticleTool')
                    if replyEl and replyEl[0].find_all('strong', class_='num'):
                        replyCount = int(replyEl[0].find_all('strong', class_='num')[0].text.strip().replace(',', ''))
                    else:
                        replyCount = '댓글수 찾기 불가'                    
                      
                    # 조회수
                    viewEl = soup.find_all(class_='article_info')
                    if viewEl and viewEl[0].find_all('span', class_='count'):
                        viewCount = int(viewEl[0].find_all('span', class_='count')[0].text.strip()[3:].replace(',', ''))
                    else:
                        viewCount = '조회수 찾기 불가'
 
                finally:
                    self.sheet[self.COLUMN_REPLY_COUNT + str(i)] = replyCount                    
                    self.sheet[self.COLUMN_VIEW_COUNT + str(i)] = viewCount
                     
            except:
                print(e)
                pass
             
            finally:
                print("댓글수:", replyCount)
                print("조회수:", viewCount)
                time.sleep(random.randrange(self.TIMEOUT_MIN, self.TIMEOUT_MAX))
         
        print('데이터 크롤링 완료')
    
    # 결과 파일 저장하기
    def saveOutputFile(self):
        print('\n출력 파일 저장')
        
        # 현재 시간
        currentTime = datetime.today().strftime("%y%m%d_%H%M%S")
        
        # 엑셀 파일 저장하기
        fileUrl = self.outputFile + "_" + currentTime + ".xlsx"
        self.book.save(fileUrl)
        
        print('출력 파일 저장 완료')
        print("결과가 \"" + fileUrl + "\"으로 저장되었습니다.")
        
    # 크롤러 실행
    def run(self):
        if not self.confirmAgreement():
            exit() 
    
        print('\n==================================================\n')
        
        while True:
            self.getInputFileName()
            if self.isExistFile(self.inputFile):
                if self.openInputFile():
                    break
        
        print('\n==================================================\n')
        
        self.getOutputFileName()    
    
        print('\n==================================================\n')
        
        while True:
            self.getUserAccount()
            
            self.connectWebDriver()
            self.login()
            if self.isLogin():
                break
    
        print('\n==================================================\n')
        
        self.openInputFile()
        self.getData()
        
        print('\n==================================================\n')
        
        self.saveOutputFile()
        self.closeWebDriver()


if __name__ == '__main__':
    crawler = Crawler()
    crawler.run()
    